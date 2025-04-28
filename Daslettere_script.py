import streamlit as st
import pandas as pd
import openpyxl
from io import StringIO
import sys

# File uploader for a single file
uploaded_file = st.file_uploader("Choose a .dat file", type=["dat"])

if uploaded_file is not None:
    # Read the dat file into a pandas DataFrame starting from the line following "Calibration"
    try:
        lines = uploaded_file.read().decode('utf-8').splitlines()
        start_line = 0
        meta_data = []
        
        for i, line in enumerate(lines):
            if "Calibration: Calibration" in line:
                start_line = i - 1  # Start reading from the line following "Calibration"
                meta_data = lines[:start_line]
                break

        # Create a DataFrame
        df = pd.read_csv(StringIO("\n".join(lines[start_line:])), sep='\t', skiprows=0)

        # Redirect print output to a buffer for later use
        buffer = StringIO()
        sys.stdout = buffer

        # Convert the DataFrame to an Excel file without writing the header
        file_name = "analyzed_data.xlsx"
        df.to_excel(file_name, index=False, header=False)

        # Open the Excel file
        workbook = openpyxl.load_workbook(file_name)
        sheet = workbook.active

        # Calculate the average and write the result to column K
        processed_strings = []
        for row in range(2, sheet.max_row + 1):
            current_value = str(sheet.cell(row, 1).value).split('_')[0]
            next_value = str(sheet.cell(row + 1, 1).value).split('_')[0] if row < sheet.max_row else ""
            if current_value != next_value:
                processed_strings.append(current_value)
                start_row = row
                end_row = row
                while end_row < sheet.max_row and str(sheet.cell(end_row + 1, 1).value).split('_')[0] == current_value:
                    end_row += 1
                if start_row != end_row:
                    for col in range(3, sheet.max_column + 1):
                        column_sum = sum(sheet.cell(r, col).value for r in range(start_row, end_row + 1) if sheet.cell(r, col).value is not None)
                        average = column_sum / (end_row - start_row + 1)
                        sheet.cell(end_row, 11, value=average)

        # Save the changes
        workbook.save(file_name)

        # Creating new sheets for metadata
        meta_sheet = workbook.create_sheet("Meta Data")
        for i, line in enumerate(meta_data):
            meta_sheet.cell(row=i + 1, column=1, value=line)

        # Create a new sheet for processed data with headers
        new_sheet = workbook.create_sheet("Processed Data")
        headers = ["label", "power", "rel mo", "abs mo", "temp set", "temp rep", "status", "date/time"]
        new_sheet.append(headers)

        # Save the final changes
        workbook.save(file_name)
        st.write("Data processing completed successfully! Download the file: ", file_name)
        
        # Optionally, you might allow users to download the processed Excel file.
        with open(file_name, "rb") as f:
            st.download_button("Download Excel file", f, file_name)

    except pd.errors.ParserError as e:
        st.error(f"Error occurred while reading the file: {e}")